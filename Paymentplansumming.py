import openpyxl
import os
import re
sumsheetrow=6
sumworkbook=openpyxl.load_workbook("汇总表.xlsx", data_only=True)
sumworksheet=sumworkbook["Sheet1"]
def plansumming(filename):
    SubWorkBook = openpyxl.load_workbook(filename, data_only=True)
    for sheetname in SubWorkBook.sheetnames:
        if SubWorkBook[sheetname].cell(3,1).value=="项目名称":#仅汇总包含项目名称的表
            sheetsumming(SubWorkBook[sheetname])
            print(sheetname)
    sumworkbook.save("汇总表.xlsx")
def sheetsumming(worksheet):#将单个工作表中的数据汇总到总表
    global sumsheetrow
    for rownum in range(22,worksheet.max_row):
        if is_text_number(worksheet.cell(rownum,1).value)==True:
            sumworksheet.cell(sumsheetrow,1).value=worksheet.cell(3,3).value
            for colnum in range(1,34):
                sumworksheet.cell(sumsheetrow,colnum+1).value=worksheet.cell(rownum,colnum).value
            sumsheetrow+=1
def is_text_number(value):
    # 使用正则表达式判断是否为文本格式的数字
    pattern = r'^[-+]?\d+(\.\d+)?$'
    return bool(re.match(pattern, str(value)))






def main():#遍历文件夹“文件源”里的文件汇总
    filenames=os.listdir(r"datasource")
    for filename in filenames:
        if filename[-5:]==".xlsx":#确保只操作xlsx文件
            plansumming(r"datasource/%s"%filename)
            print(filename)
main()